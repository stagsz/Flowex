"""Data list export service for P&ID drawings (Excel, CSV, PDF)."""

import csv
import logging
import tempfile
from collections import Counter
from dataclasses import dataclass, field
from datetime import UTC, datetime
from enum import Enum
from pathlib import Path
from typing import TYPE_CHECKING, Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

if TYPE_CHECKING:
    from app.models.drawing import Drawing
    from app.models.line import Line
    from app.models.symbol import Symbol
    from app.models.text_annotation import TextAnnotation

logger = logging.getLogger(__name__)


class ExportFormat(str, Enum):
    """Supported export formats for data lists."""

    XLSX = "xlsx"
    CSV = "csv"
    PDF = "pdf"


class ListType(str, Enum):
    """Types of data lists that can be exported."""

    EQUIPMENT = "equipment"
    LINE = "line"
    INSTRUMENT = "instrument"
    VALVE = "valve"
    MTO = "mto"
    CHECKLIST = "checklist"


@dataclass
class ExportMetadata:
    """Metadata for exported data lists."""

    project_name: str
    drawing_number: str
    revision: str = "A"
    date: str = field(default_factory=lambda: datetime.now(UTC).strftime("%Y-%m-%d"))
    prepared_by: str = "Flowex"


# Excel styling constants
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
ALT_ROW_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


class DataListExportService:
    """Service for exporting engineering data lists."""

    def __init__(self) -> None:
        self.output_dir = Path(tempfile.gettempdir()) / "flowex_exports"
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def export_equipment_list(
        self,
        drawing: "Drawing",
        symbols: list["Symbol"],
        metadata: ExportMetadata,
        format: ExportFormat = ExportFormat.XLSX,
        include_unverified: bool = False,
    ) -> Path:
        """Export equipment list."""
        # Filter equipment symbols
        equipment = [
            s for s in symbols
            if s.category.value == "equipment"
            and not s.is_deleted
            and (include_unverified or s.is_verified)
        ]

        headers = [
            "Tag Number",
            "Description",
            "Type",
            "Size/Capacity",
            "Material",
            "Design Pressure",
            "Design Temperature",
            "Drawing Reference",
            "Notes",
        ]

        rows = []
        for eq in equipment:
            rows.append([
                eq.tag_number or "-",
                self._get_description_from_class(eq.symbol_class),
                eq.symbol_class.replace("_", " "),
                "-",  # Size/Capacity - would come from additional data
                "-",  # Material
                "-",  # Design Pressure
                "-",  # Design Temperature
                metadata.drawing_number,
                "Verified" if eq.is_verified else "Unverified",
            ])

        return self._export_list(
            headers, rows, metadata, "Equipment_List", format
        )

    def export_line_list(
        self,
        drawing: "Drawing",
        lines: list["Line"],
        metadata: ExportMetadata,
        format: ExportFormat = ExportFormat.XLSX,
        include_unverified: bool = False,
    ) -> Path:
        """Export line list."""
        filtered_lines = [
            ln for ln in lines
            if not ln.is_deleted
            and (include_unverified or ln.is_verified)
        ]

        headers = [
            "Line Number",
            "Size",
            "Spec",
            "From",
            "To",
            "Fluid",
            "Design Pressure",
            "Design Temperature",
            "Insulation",
            "Material",
            "Drawing Reference",
            "Notes",
        ]

        rows = []
        for ln in filtered_lines:
            # Parse line spec if available (e.g., "6"-P-101-A1")
            size = self._parse_line_size(ln.line_spec)
            spec = ln.pipe_class or "-"

            rows.append([
                ln.line_number or "-",
                size,
                spec,
                "-",  # From - would need connectivity data
                "-",  # To
                "-",  # Fluid
                "-",  # Design Pressure
                "-",  # Design Temperature
                ln.insulation or "-",
                "-",  # Material
                metadata.drawing_number,
                "Verified" if ln.is_verified else "Unverified",
            ])

        return self._export_list(
            headers, rows, metadata, "Line_List", format
        )

    def export_instrument_list(
        self,
        drawing: "Drawing",
        symbols: list["Symbol"],
        metadata: ExportMetadata,
        format: ExportFormat = ExportFormat.XLSX,
        include_unverified: bool = False,
    ) -> Path:
        """Export instrument list."""
        instruments = [
            s for s in symbols
            if s.category.value == "instrument"
            and not s.is_deleted
            and (include_unverified or s.is_verified)
        ]

        headers = [
            "Tag Number",
            "Type",
            "Description",
            "Service",
            "Range",
            "Units",
            "Output",
            "Location",
            "Drawing Reference",
            "Loop Number",
            "Notes",
        ]

        rows = []
        for inst in instruments:
            inst_type = self._get_instrument_type(inst.symbol_class)

            rows.append([
                inst.tag_number or "-",
                inst_type,
                self._get_description_from_class(inst.symbol_class),
                "-",  # Service
                "-",  # Range
                "-",  # Units
                "-",  # Output
                "Field",  # Location
                metadata.drawing_number,
                self._extract_loop_number(inst.tag_number),
                "Verified" if inst.is_verified else "Unverified",
            ])

        return self._export_list(
            headers, rows, metadata, "Instrument_List", format
        )

    def export_valve_list(
        self,
        drawing: "Drawing",
        symbols: list["Symbol"],
        metadata: ExportMetadata,
        format: ExportFormat = ExportFormat.XLSX,
        include_unverified: bool = False,
    ) -> Path:
        """Export valve list."""
        valves = [
            s for s in symbols
            if s.category.value == "valve"
            and not s.is_deleted
            and (include_unverified or s.is_verified)
        ]

        headers = [
            "Tag Number",
            "Type",
            "Size",
            "Rating",
            "Spec",
            "Line Number",
            "Body Material",
            "Trim Material",
            "Actuator",
            "Fail Position",
            "Drawing Reference",
            "Notes",
        ]

        rows = []
        for valve in valves:
            valve_type = self._get_valve_type(valve.symbol_class)
            actuator = self._get_actuator_type(valve.symbol_class)

            rows.append([
                valve.tag_number or "-",
                valve_type,
                "-",  # Size
                "-",  # Rating
                "-",  # Spec
                "-",  # Line Number
                "-",  # Body Material
                "-",  # Trim Material
                actuator,
                "-",  # Fail Position
                metadata.drawing_number,
                "Verified" if valve.is_verified else "Unverified",
            ])

        return self._export_list(
            headers, rows, metadata, "Valve_List", format
        )

    def export_mto(
        self,
        drawing: "Drawing",
        symbols: list["Symbol"],
        lines: list["Line"],
        metadata: ExportMetadata,
        format: ExportFormat = ExportFormat.XLSX,
        include_unverified: bool = False,
    ) -> Path:
        """Export Material Take-Off (MTO)."""
        headers = [
            "Item Number",
            "Category",
            "Description",
            "Size",
            "Material",
            "Quantity",
            "Unit",
            "Spec",
            "Drawing Reference",
            "Notes",
        ]

        # Count items by type
        item_counts: Counter[str] = Counter()

        # Count symbols
        for symbol in symbols:
            if symbol.is_deleted:
                continue
            if not include_unverified and not symbol.is_verified:
                continue
            item_counts[symbol.symbol_class] += 1

        # Build rows
        rows = []
        item_num = 1

        for symbol_class, count in sorted(item_counts.items()):
            category = self._get_category_for_class(symbol_class)
            description = self._get_description_from_class(symbol_class)

            rows.append([
                item_num,
                category,
                description,
                "-",  # Size
                "-",  # Material
                count,
                "EA",  # Unit
                "-",  # Spec
                metadata.drawing_number,
                "",
            ])
            item_num += 1

        # Count lines by type
        line_counts: Counter[str] = Counter()
        for line in lines:
            if line.is_deleted:
                continue
            if not include_unverified and not line.is_verified:
                continue
            line_type = line.pipe_class or "Unknown"
            line_counts[line_type] += 1

        for line_type, count in sorted(line_counts.items()):
            rows.append([
                item_num,
                "Piping",
                f"Pipe - {line_type}",
                "-",
                "-",
                count,
                "M",  # Meters
                line_type,
                metadata.drawing_number,
                "",
            ])
            item_num += 1

        return self._export_list(
            headers, rows, metadata, "MTO", format
        )

    def export_comparison_report(
        self,
        drawing: "Drawing",
        symbols: list["Symbol"],
        lines: list["Line"],
        text_annotations: list["TextAnnotation"],
        metadata: ExportMetadata,
        format: ExportFormat = ExportFormat.PDF,
    ) -> Path:
        """Export extraction summary/comparison report."""
        # Gather statistics
        stats = self._calculate_statistics(symbols, lines, text_annotations)
        flagged_items = self._get_flagged_items(symbols, lines, text_annotations)

        if format == ExportFormat.PDF:
            return self._export_report_pdf(
                drawing, stats, flagged_items, metadata
            )
        else:
            # For Excel/CSV, export as simple list
            return self._export_report_tabular(
                drawing, stats, flagged_items, metadata, format
            )

    def export_validation_checklist(
        self,
        drawing: "Drawing",
        symbols: list["Symbol"],
        lines: list["Line"],
        metadata: ExportMetadata,
        format: ExportFormat = ExportFormat.PDF,
        include_unverified: bool = True,
    ) -> Path:
        """
        Export validation checklist with all items and their verification status.

        This is the item-by-item checklist that allows users to track what
        has been reviewed during the validation process.
        """
        if format == ExportFormat.PDF:
            return self._export_checklist_pdf(
                drawing, symbols, lines, metadata, include_unverified
            )
        else:
            return self._export_checklist_tabular(
                drawing, symbols, lines, metadata, format, include_unverified
            )

    def _export_list(
        self,
        headers: list[str],
        rows: list[list[Any]],
        metadata: ExportMetadata,
        list_name: str,
        format: ExportFormat,
    ) -> Path:
        """Export a data list in the specified format."""
        if format == ExportFormat.XLSX:
            return self._export_xlsx(headers, rows, metadata, list_name)
        elif format == ExportFormat.CSV:
            return self._export_csv(headers, rows, metadata, list_name)
        elif format == ExportFormat.PDF:
            return self._export_pdf(headers, rows, metadata, list_name)
        else:
            raise ValueError(f"Unsupported format: {format}")

    def _export_xlsx(
        self,
        headers: list[str],
        rows: list[list[Any]],
        metadata: ExportMetadata,
        list_name: str,
    ) -> Path:
        """Export to Excel format."""
        wb = Workbook()
        ws = wb.active
        ws.title = list_name.replace("_", " ")

        # Add metadata header
        ws.merge_cells("A1:D1")
        ws["A1"] = list_name.replace("_", " ").upper()
        ws["A1"].font = Font(bold=True, size=14)

        ws["A2"] = f"Project: {metadata.project_name}"
        ws["C2"] = f"Date: {metadata.date}"
        ws["A3"] = f"Drawing: {metadata.drawing_number}"
        ws["C3"] = f"Rev: {metadata.revision}"

        # Add column headers (row 5)
        header_row = 5
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Add data rows
        for row_idx, row_data in enumerate(rows, header_row + 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = THIN_BORDER
                cell.alignment = Alignment(vertical="center")

                # Alternating row colors
                if (row_idx - header_row) % 2 == 0:
                    cell.fill = ALT_ROW_FILL

        # Auto-fit column widths
        for col in range(1, len(headers) + 1):
            max_length = len(str(headers[col - 1]))
            for row in rows:
                if col - 1 < len(row):
                    max_length = max(max_length, len(str(row[col - 1])))
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[get_column_letter(col)].width = adjusted_width

        # Enable filters
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{header_row + len(rows)}"

        # Freeze header row
        ws.freeze_panes = f"A{header_row + 1}"

        # Save file
        filename = f"{metadata.drawing_number}_{list_name}.xlsx"
        output_path = self.output_dir / filename
        wb.save(output_path)

        logger.info(f"Exported Excel file: {output_path}")
        return output_path

    def _export_csv(
        self,
        headers: list[str],
        rows: list[list[Any]],
        metadata: ExportMetadata,
        list_name: str,
    ) -> Path:
        """Export to CSV format."""
        filename = f"{metadata.drawing_number}_{list_name}.csv"
        output_path = self.output_dir / filename

        with open(output_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)

            # Write metadata as comments
            writer.writerow([f"# {list_name.replace('_', ' ')}"])
            writer.writerow([f"# Project: {metadata.project_name}"])
            writer.writerow([f"# Drawing: {metadata.drawing_number}"])
            writer.writerow([f"# Date: {metadata.date}"])
            writer.writerow([])

            # Write headers and data
            writer.writerow(headers)
            writer.writerows(rows)

        logger.info(f"Exported CSV file: {output_path}")
        return output_path

    def _export_pdf(
        self,
        headers: list[str],
        rows: list[list[Any]],
        metadata: ExportMetadata,
        list_name: str,
    ) -> Path:
        """Export to PDF format."""
        filename = f"{metadata.drawing_number}_{list_name}.pdf"
        output_path = self.output_dir / filename

        doc = SimpleDocTemplate(
            str(output_path),
            pagesize=landscape(A4),
            rightMargin=20 * mm,
            leftMargin=20 * mm,
            topMargin=20 * mm,
            bottomMargin=20 * mm,
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "Title",
            parent=styles["Heading1"],
            fontSize=16,
            alignment=1,  # Center
        )

        elements = []

        # Title
        elements.append(Paragraph(list_name.replace("_", " ").upper(), title_style))
        elements.append(Spacer(1, 10 * mm))

        # Metadata
        meta_text = (
            f"Project: {metadata.project_name} | "
            f"Drawing: {metadata.drawing_number} | "
            f"Rev: {metadata.revision} | "
            f"Date: {metadata.date}"
        )
        elements.append(Paragraph(meta_text, styles["Normal"]))
        elements.append(Spacer(1, 10 * mm))

        # Table
        table_data = [headers] + [[str(v) for v in row] for row in rows]

        # Calculate column widths
        available_width = landscape(A4)[0] - 40 * mm
        col_width = available_width / len(headers)
        col_widths = [col_width] * len(headers)

        table = Table(table_data, colWidths=col_widths, repeatRows=1)
        table.setStyle(TableStyle([
            # Header style
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("ALIGN", (0, 0), (-1, 0), "CENTER"),

            # Data rows
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 1), (-1, -1), 8),
            ("ALIGN", (0, 1), (-1, -1), "LEFT"),

            # Grid
            ("GRID", (0, 0), (-1, -1), 0.5, colors.black),

            # Alternating row colors
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#D9E2F3")]),

            # Padding
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))

        elements.append(table)

        # Summary
        elements.append(Spacer(1, 10 * mm))
        elements.append(Paragraph(f"Total Items: {len(rows)}", styles["Normal"]))

        # Build PDF
        doc.build(elements)

        logger.info(f"Exported PDF file: {output_path}")
        return output_path

    def _calculate_statistics(
        self,
        symbols: list["Symbol"],
        lines: list["Line"],
        text_annotations: list["TextAnnotation"],
    ) -> dict[str, dict[str, int]]:
        """Calculate extraction statistics."""
        stats = {
            "Equipment": {"count": 0, "verified": 0, "flagged": 0},
            "Instruments": {"count": 0, "verified": 0, "flagged": 0},
            "Valves": {"count": 0, "verified": 0, "flagged": 0},
            "Lines": {"count": 0, "verified": 0, "flagged": 0},
        }

        for symbol in symbols:
            if symbol.is_deleted:
                continue

            category = symbol.category.value
            if category == "equipment":
                key = "Equipment"
            elif category == "instrument":
                key = "Instruments"
            elif category == "valve":
                key = "Valves"
            else:
                key = "Equipment"

            stats[key]["count"] += 1
            if symbol.is_verified:
                stats[key]["verified"] += 1
            if symbol.confidence and symbol.confidence < 0.7:
                stats[key]["flagged"] += 1

        for line in lines:
            if line.is_deleted:
                continue
            stats["Lines"]["count"] += 1
            if line.is_verified:
                stats["Lines"]["verified"] += 1
            if line.confidence and line.confidence < 0.7:
                stats["Lines"]["flagged"] += 1

        return stats

    def _get_flagged_items(
        self,
        symbols: list["Symbol"],
        lines: list["Line"],
        text_annotations: list["TextAnnotation"],
    ) -> list[dict[str, Any]]:
        """Get list of items flagged for review."""
        flagged = []

        for symbol in symbols:
            if symbol.is_deleted:
                continue
            if symbol.confidence and symbol.confidence < 0.7:
                flagged.append({
                    "tag": symbol.tag_number or symbol.symbol_class,
                    "type": "Symbol",
                    "reason": f"Low confidence ({symbol.confidence:.2f})",
                })

        for text in text_annotations:
            if text.is_deleted:
                continue
            if text.confidence and text.confidence < 0.7:
                flagged.append({
                    "tag": text.text_content[:20],
                    "type": "Text",
                    "reason": f"OCR confidence ({text.confidence:.2f})",
                })

        return flagged

    def _export_report_pdf(
        self,
        drawing: "Drawing",
        stats: dict[str, dict[str, int]],
        flagged_items: list[dict[str, Any]],
        metadata: ExportMetadata,
    ) -> Path:
        """Export comparison report as PDF."""
        filename = f"{metadata.drawing_number}_Comparison_Report.pdf"
        output_path = self.output_dir / filename

        doc = SimpleDocTemplate(
            str(output_path),
            pagesize=A4,
            rightMargin=20 * mm,
            leftMargin=20 * mm,
            topMargin=20 * mm,
            bottomMargin=20 * mm,
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "Title",
            parent=styles["Heading1"],
            fontSize=18,
            alignment=1,
        )
        section_style = ParagraphStyle(
            "Section",
            parent=styles["Heading2"],
            fontSize=12,
            spaceAfter=6 * mm,
        )

        elements = []

        # Title
        elements.append(Paragraph("EXTRACTION SUMMARY REPORT", title_style))
        elements.append(Spacer(1, 10 * mm))

        # 1. Drawing Information
        elements.append(Paragraph("1. DRAWING INFORMATION", section_style))
        info_data = [
            ["Drawing Number:", metadata.drawing_number],
            ["Revision:", metadata.revision],
            ["Project:", metadata.project_name],
            ["Processed Date:", metadata.date],
            ["Processed By:", metadata.prepared_by],
        ]
        info_table = Table(info_data, colWidths=[50 * mm, 100 * mm])
        info_table.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
        ]))
        elements.append(info_table)
        elements.append(Spacer(1, 10 * mm))

        # 2. Extraction Statistics
        elements.append(Paragraph("2. EXTRACTION STATISTICS", section_style))

        stat_headers = ["Category", "Count", "Verified", "Flagged"]
        stat_rows = []
        total_count = total_verified = total_flagged = 0

        for category, values in stats.items():
            stat_rows.append([
                category,
                str(values["count"]),
                str(values["verified"]),
                str(values["flagged"]),
            ])
            total_count += values["count"]
            total_verified += values["verified"]
            total_flagged += values["flagged"]

        stat_rows.append(["TOTAL", str(total_count), str(total_verified), str(total_flagged)])

        stat_table = Table(
            [stat_headers] + stat_rows,
            colWidths=[50 * mm, 30 * mm, 30 * mm, 30 * mm],
        )
        stat_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
            ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#D9E2F3")),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
            ("ALIGN", (1, 0), (-1, -1), "CENTER"),
        ]))
        elements.append(stat_table)
        elements.append(Spacer(1, 10 * mm))

        # 3. Items Flagged for Review
        elements.append(Paragraph("3. ITEMS FLAGGED FOR REVIEW", section_style))

        if flagged_items:
            flagged_headers = ["Item", "Type", "Reason"]
            flagged_rows = [[f["tag"], f["type"], f["reason"]] for f in flagged_items[:20]]

            flagged_table = Table(
                [flagged_headers] + flagged_rows,
                colWidths=[50 * mm, 30 * mm, 70 * mm],
            )
            flagged_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#C55A11")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
            ]))
            elements.append(flagged_table)

            if len(flagged_items) > 20:
                elements.append(Paragraph(
                    f"... and {len(flagged_items) - 20} more items",
                    styles["Italic"],
                ))
        else:
            elements.append(Paragraph("No items flagged for review.", styles["Normal"]))

        elements.append(Spacer(1, 10 * mm))

        # 4. Validation Status
        elements.append(Paragraph("4. VALIDATION STATUS", section_style))
        if total_count > 0:
            completion = (total_verified / total_count) * 100
            elements.append(Paragraph(
                f"Completed: {total_verified}/{total_count} ({completion:.1f}%)",
                styles["Normal"],
            ))
        else:
            elements.append(Paragraph("No items to validate.", styles["Normal"]))

        # Build PDF
        doc.build(elements)

        logger.info(f"Exported comparison report: {output_path}")
        return output_path

    def _export_report_tabular(
        self,
        drawing: "Drawing",
        stats: dict[str, dict[str, int]],
        flagged_items: list[dict[str, Any]],
        metadata: ExportMetadata,
        format: ExportFormat,
    ) -> Path:
        """Export comparison report as tabular format (Excel/CSV)."""
        headers = ["Category", "Count", "Verified", "Flagged"]
        rows = []

        for category, values in stats.items():
            rows.append([
                category,
                values["count"],
                values["verified"],
                values["flagged"],
            ])

        return self._export_list(
            headers, rows, metadata, "Comparison_Report", format
        )

    def _export_checklist_pdf(
        self,
        drawing: "Drawing",
        symbols: list["Symbol"],
        lines: list["Line"],
        metadata: ExportMetadata,
        include_unverified: bool,
    ) -> Path:
        """Export validation checklist as PDF."""
        filename = f"{metadata.drawing_number}_Validation_Checklist.pdf"
        output_path = self.output_dir / filename

        doc = SimpleDocTemplate(
            str(output_path),
            pagesize=A4,
            rightMargin=15 * mm,
            leftMargin=15 * mm,
            topMargin=15 * mm,
            bottomMargin=15 * mm,
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "Title",
            parent=styles["Heading1"],
            fontSize=16,
            alignment=1,
        )
        section_style = ParagraphStyle(
            "Section",
            parent=styles["Heading2"],
            fontSize=11,
            spaceAfter=4 * mm,
            spaceBefore=6 * mm,
        )

        elements = []

        # Title
        elements.append(Paragraph("VALIDATION CHECKLIST", title_style))
        elements.append(Spacer(1, 6 * mm))

        # Drawing Information
        info_data = [
            ["Drawing:", metadata.drawing_number, "Project:", metadata.project_name],
            ["Revision:", metadata.revision, "Date:", metadata.date],
        ]
        info_table = Table(info_data, colWidths=[25 * mm, 55 * mm, 25 * mm, 55 * mm])
        info_table.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
            ("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ]))
        elements.append(info_table)
        elements.append(Spacer(1, 6 * mm))

        # Group symbols by category
        categories: dict[str, list[Symbol]] = {
            "Equipment": [],
            "Instruments": [],
            "Valves": [],
            "Other": [],
        }

        for symbol in symbols:
            if symbol.is_deleted:
                continue
            if not include_unverified and not symbol.is_verified:
                continue

            category = symbol.category.value
            if category == "equipment":
                categories["Equipment"].append(symbol)
            elif category == "instrument":
                categories["Instruments"].append(symbol)
            elif category == "valve":
                categories["Valves"].append(symbol)
            else:
                categories["Other"].append(symbol)

        # Track totals
        total_items = 0
        verified_items = 0
        flagged_items = 0

        # Export each category
        for cat_name, cat_symbols in categories.items():
            if not cat_symbols:
                continue

            elements.append(Paragraph(cat_name.upper(), section_style))

            headers = ["Status", "Tag Number", "Type", "Confidence", "Flagged", "Notes"]
            rows = []

            for symbol in sorted(cat_symbols, key=lambda s: s.tag_number or ""):
                status = "✓" if symbol.is_verified else "○"
                is_flagged = getattr(symbol, "is_flagged", False) or (
                    symbol.confidence and symbol.confidence < 0.7
                )
                flag_indicator = "⚠" if is_flagged else ""
                confidence = f"{symbol.confidence:.0%}" if symbol.confidence else "-"

                rows.append([
                    status,
                    symbol.tag_number or "-",
                    self._get_description_from_class(symbol.symbol_class),
                    confidence,
                    flag_indicator,
                    "",  # Empty notes column for manual entry
                ])

                total_items += 1
                if symbol.is_verified:
                    verified_items += 1
                if is_flagged:
                    flagged_items += 1

            # Create table with appropriate column widths
            col_widths = [12 * mm, 30 * mm, 55 * mm, 22 * mm, 15 * mm, 40 * mm]
            table = Table([headers] + rows, colWidths=col_widths, repeatRows=1)
            table.setStyle(TableStyle([
                # Header
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 8),
                ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                # Data
                ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                ("FONTSIZE", (0, 1), (-1, -1), 8),
                ("ALIGN", (0, 1), (0, -1), "CENTER"),  # Status column centered
                ("ALIGN", (3, 1), (4, -1), "CENTER"),  # Confidence and Flag centered
                # Grid
                ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                # Row backgrounds
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
                # Padding
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]))
            elements.append(table)

        # Lines section
        active_lines = [ln for ln in lines if not ln.is_deleted and (include_unverified or ln.is_verified)]
        if active_lines:
            elements.append(Paragraph("LINES", section_style))

            headers = ["Status", "Line Number", "Spec", "Pipe Class", "Confidence", "Flagged", "Notes"]
            rows = []

            for line in sorted(active_lines, key=lambda ln: ln.line_number or ""):
                status = "✓" if line.is_verified else "○"
                is_flagged = getattr(line, "is_flagged", False) or (
                    line.confidence and line.confidence < 0.7
                )
                flag_indicator = "⚠" if is_flagged else ""
                confidence = f"{line.confidence:.0%}" if line.confidence else "-"

                rows.append([
                    status,
                    line.line_number or "-",
                    line.line_spec or "-",
                    line.pipe_class or "-",
                    confidence,
                    flag_indicator,
                    "",
                ])

                total_items += 1
                if line.is_verified:
                    verified_items += 1
                if is_flagged:
                    flagged_items += 1

            col_widths = [12 * mm, 30 * mm, 35 * mm, 25 * mm, 20 * mm, 15 * mm, 35 * mm]
            table = Table([headers] + rows, colWidths=col_widths, repeatRows=1)
            table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 8),
                ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                ("FONTSIZE", (0, 1), (-1, -1), 8),
                ("ALIGN", (0, 1), (0, -1), "CENTER"),
                ("ALIGN", (4, 1), (5, -1), "CENTER"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]))
            elements.append(table)

        # Summary section
        elements.append(Spacer(1, 10 * mm))
        elements.append(Paragraph("SUMMARY", section_style))

        completion_pct = (verified_items / total_items * 100) if total_items > 0 else 0
        summary_data = [
            ["Total Items:", str(total_items)],
            ["Verified:", f"{verified_items} ({completion_pct:.1f}%)"],
            ["Pending:", str(total_items - verified_items)],
            ["Flagged for Review:", str(flagged_items)],
        ]
        summary_table = Table(summary_data, colWidths=[45 * mm, 40 * mm])
        summary_table.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))
        elements.append(summary_table)

        # Legend
        elements.append(Spacer(1, 8 * mm))
        legend_text = "Legend: ✓ = Verified, ○ = Pending, ⚠ = Flagged for Review"
        elements.append(Paragraph(legend_text, styles["Italic"]))

        # Signature section
        elements.append(Spacer(1, 15 * mm))
        sig_data = [
            ["Validated By:", "_" * 30, "Date:", "_" * 20],
            ["Approved By:", "_" * 30, "Date:", "_" * 20],
        ]
        sig_table = Table(sig_data, colWidths=[30 * mm, 65 * mm, 20 * mm, 45 * mm])
        sig_table.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("TOPPADDING", (0, 0), (-1, -1), 8),
        ]))
        elements.append(sig_table)

        # Build PDF
        doc.build(elements)

        logger.info(f"Exported validation checklist: {output_path}")
        return output_path

    def _export_checklist_tabular(
        self,
        drawing: "Drawing",
        symbols: list["Symbol"],
        lines: list["Line"],
        metadata: ExportMetadata,
        format: ExportFormat,
        include_unverified: bool,
    ) -> Path:
        """Export validation checklist as Excel/CSV format."""
        headers = [
            "Category",
            "Tag Number",
            "Type",
            "Status",
            "Confidence",
            "Flagged",
            "Drawing Reference",
        ]
        rows = []

        # Process symbols
        for symbol in symbols:
            if symbol.is_deleted:
                continue
            if not include_unverified and not symbol.is_verified:
                continue

            category = symbol.category.value.title()
            status = "Verified" if symbol.is_verified else "Pending"
            is_flagged = getattr(symbol, "is_flagged", False) or (
                symbol.confidence and symbol.confidence < 0.7
            )
            confidence = f"{symbol.confidence:.0%}" if symbol.confidence else "-"

            rows.append([
                category,
                symbol.tag_number or "-",
                self._get_description_from_class(symbol.symbol_class),
                status,
                confidence,
                "Yes" if is_flagged else "No",
                metadata.drawing_number,
            ])

        # Process lines
        for line in lines:
            if line.is_deleted:
                continue
            if not include_unverified and not line.is_verified:
                continue

            status = "Verified" if line.is_verified else "Pending"
            is_flagged = getattr(line, "is_flagged", False) or (
                line.confidence and line.confidence < 0.7
            )
            confidence = f"{line.confidence:.0%}" if line.confidence else "-"

            rows.append([
                "Line",
                line.line_number or "-",
                line.pipe_class or "-",
                status,
                confidence,
                "Yes" if is_flagged else "No",
                metadata.drawing_number,
            ])

        return self._export_list(
            headers, rows, metadata, "Validation_Checklist", format
        )

    # Helper methods

    def _get_description_from_class(self, symbol_class: str) -> str:
        """Convert symbol class name to readable description."""
        return symbol_class.replace("_", " ").title()

    def _parse_line_size(self, line_spec: str | None) -> str:
        """Extract pipe size from line specification."""
        if not line_spec:
            return "-"
        # Line spec format: "6"-P-101-A1" - size is first part
        parts = line_spec.split("-")
        if parts and parts[0].replace('"', "").replace("'", "").strip():
            return parts[0].strip()
        return "-"

    def _get_instrument_type(self, symbol_class: str) -> str:
        """Extract instrument type from symbol class."""
        type_mapping = {
            "Transmitter_Pressure": "PT",
            "Transmitter_Temperature": "TT",
            "Transmitter_Flow": "FT",
            "Transmitter_Level": "LT",
            "Controller_Generic": "C",
            "Indicator_Generic": "I",
            "Alarm_High": "AH",
            "Alarm_Low": "AL",
            "Switch_Generic": "S",
            "Control_Valve_Globe": "CV",
            "Control_Valve_Butterfly": "CV",
            "Orifice_Plate": "FE",
            "Thermowell": "TW",
            "Sample_Point": "SP",
            "Relief_Valve_Instrument": "PSV",
        }
        return type_mapping.get(symbol_class, "I")

    def _get_valve_type(self, symbol_class: str) -> str:
        """Extract valve type from symbol class."""
        type_mapping = {
            "Valve_Gate": "Gate",
            "Valve_Globe": "Globe",
            "Valve_Ball": "Ball",
            "Valve_Butterfly": "Butterfly",
            "Valve_Check": "Check",
            "Valve_Relief_PSV": "Relief/Safety",
            "Valve_Control": "Control",
            "Valve_Three_Way": "3-Way",
            "Valve_Diaphragm": "Diaphragm",
            "Valve_Plug": "Plug",
            "Valve_Needle": "Needle",
            "Valve_Manual_Generic": "Manual",
        }
        return type_mapping.get(symbol_class, "Manual")

    def _get_actuator_type(self, symbol_class: str) -> str:
        """Determine actuator type from symbol class."""
        if "Pneumatic" in symbol_class:
            return "Pneumatic"
        elif "Electric" in symbol_class:
            return "Electric"
        elif "Hydraulic" in symbol_class:
            return "Hydraulic"
        elif "Control" in symbol_class:
            return "Pneumatic"  # Default for control valves
        return "Manual"

    def _extract_loop_number(self, tag_number: str | None) -> str:
        """Extract loop number from instrument tag."""
        if not tag_number:
            return "-"
        # Tag format: FT-101, extract 101 as loop
        import re
        match = re.search(r"\d+", tag_number)
        return match.group() if match else "-"

    def _get_category_for_class(self, symbol_class: str) -> str:
        """Get category name for a symbol class."""
        if any(x in symbol_class for x in ["Vessel", "Tank", "Pump", "Compressor",
                                            "Heat_Exchanger", "Column", "Filter",
                                            "Reactor", "Furnace", "Blower", "Agitator"]):
            return "Equipment"
        elif any(x in symbol_class for x in ["Transmitter", "Controller", "Indicator",
                                              "Alarm", "Switch", "Orifice", "Thermowell",
                                              "Sample", "Relief_Valve_Instrument"]):
            return "Instruments"
        elif any(x in symbol_class for x in ["Valve", "Actuator"]):
            return "Valves"
        else:
            return "Other"


def export_equipment_list(
    drawing: "Drawing",
    symbols: list["Symbol"],
    metadata: ExportMetadata,
    format: ExportFormat = ExportFormat.XLSX,
) -> Path:
    """Convenience function to export equipment list."""
    service = DataListExportService()
    return service.export_equipment_list(drawing, symbols, metadata, format)


def export_all_lists(
    drawing: "Drawing",
    symbols: list["Symbol"],
    lines: list["Line"],
    text_annotations: list["TextAnnotation"],
    metadata: ExportMetadata,
    format: ExportFormat = ExportFormat.XLSX,
    include_unverified: bool = False,
) -> dict[str, Path]:
    """Export all data lists for a drawing."""
    service = DataListExportService()

    return {
        "equipment": service.export_equipment_list(
            drawing, symbols, metadata, format, include_unverified
        ),
        "line": service.export_line_list(
            drawing, lines, metadata, format, include_unverified
        ),
        "instrument": service.export_instrument_list(
            drawing, symbols, metadata, format, include_unverified
        ),
        "valve": service.export_valve_list(
            drawing, symbols, metadata, format, include_unverified
        ),
        "mto": service.export_mto(
            drawing, symbols, lines, metadata, format, include_unverified
        ),
        "report": service.export_comparison_report(
            drawing, symbols, lines, text_annotations, metadata, format
        ),
    }
